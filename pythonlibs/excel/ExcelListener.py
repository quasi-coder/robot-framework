import os
import openpyxl
from robot.libraries.BuiltIn import BuiltIn
from openpyxl.styles.fills import PatternFill

class ExcelListener(object):

    ROBOT_LISTENER_API_VERSION = 2

    _builtin = BuiltIn()

    def __init__(self):
        self._workbook = None
        self._keyword_message = ''
        self._level = 0
        self._test_steps = {}
        self._current_test_name = ''
        self._current_suite_name = ''

    @property
    def _robot_vars(self):
        '''
        Gets robot variables.
        '''
        return self._builtin.get_variables()

    def __add_step(self, name, attrs):
        if (attrs['type'] in ('Suite Setup', 'Suite Teardown', 'Test Setup', 'Test Teardown', 'Keyword') and self._level == 1) or \
                (attrs['type'] in ('Test Foritem') and self._level == 2):
            if attrs['type'] in ('Keyword', 'Test Foritem'):
                key = self._current_test_name
            else:
                key = attrs['type']

            if attrs['type'] == 'Test Foritem':
                longname = self._robot_vars.get('${' + self._current_test_name + '_SCENARIO_NAME}')
                expected = self._robot_vars.get('${' + self._current_test_name + '_EXPECTED_RESULT}')
            else:
                longname = name
                expected = 'PASS'

            if not longname:
                logger.warning('Can not obtain long name for keyword "%s"', name)
                return
            test_step = {
                'template': attrs.get('template', ''),
                'expected': expected,
                'longname': longname,
                'name': key,
                'params': ', '.join(attrs['args']),
                'status': attrs['status'],
                'message': self._keyword_message,
                'tags': self._robot_vars.get('@{TEST_TAGS}', []),
                'doc': attrs['doc'],
            }
            # If we are on Suite Setup stage, there is no self._test_steps[key] is defined yet.
            if key not in self._test_steps:
                self._test_steps[key] = []
            self._test_steps[key].append(test_step)
        elif attrs['type'] == 'Test Teardown' and name == 'util.Report Status Message':
            # Save keyword status message for later use when process the for item
            # 'util.Report Status Message' keyword should be called from test template keyword teardown to enable this
            self._keyword_message = self._robot_vars.get('${KEYWORD_MESSAGE}')

    def start_test(self, name, attrs):
        self._current_test_name = name
        if not self._workbook:
            if os.path.exists(self.get_output_path()):
                self._workbook = openpyxl.load_workbook(self.get_output_path(), data_only=True)
            else:
                self._workbook = openpyxl.Workbook()

        if name in self._workbook:
            self._workbook.remove_sheet(self._workbook[name])

        self._workbook.create_sheet(0, name)
        self._workbook.active.append(['Scenario', 'Expected', 'Status', 'Message'])
        self.set_last_row_color('C6EFCE')

        self._test_steps[name] = []

    def end_test(self, name, attrs):
        self._level = 0
        self._test_steps[name] = self._test_steps.pop('Test Setup', []) + self._test_steps[name]
        self._test_steps[name] = self._test_steps.get('Suite Setup', []) + self._test_steps[name]
        self._test_steps[name] += self._test_steps.pop('Test Teardown', [])
        self._test_steps[name] += self._test_steps.get('Suite Teardown', [])
        self._current_test_name = ''
        for step in self._test_steps[name]:
            self._workbook.active.append([step['longname'], step['expected'], step['status'], step['message']])
            if step['status'] == 'FAIL':
                self.set_last_row_color('00FFC7CE')

    def start_keyword(self, name, attrs):
        self._level += 1

    def end_keyword(self, name, attrs):
        try:
            self.__add_step(name, attrs)
        finally:
            self._level -= 1

    def start_suite(self, name, attrs):
        self._current_suite_name = name

    def end_suite(self, name, attrs):
        try:
            if self._workbook:
                self._workbook.save(self.get_output_path())
        finally:
            self._test_steps = {}
            self._workbook = None
            self._current_suite_name = ''

    def get_output_path(self):
        output_dir = self._robot_vars.get('${OUTPUT_DIR}') + '/excel'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        output_path = output_dir + '/' + self._robot_vars.get('${SUITE_NAME}') + '.xlsx'
        return output_path

    def set_last_row_color(self, color):
        max_row = self._workbook.active.max_row
        fill = PatternFill(patternType='solid', fgColor=color)
        self._workbook.active.row_dimensions[max_row].fill = fill
        for cell in self._workbook.active.rows[max_row - 1]:
            cell.fill = PatternFill(patternType='solid', fgColor=color)
